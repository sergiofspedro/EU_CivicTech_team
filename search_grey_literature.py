"""
search_grey_literature.py
Searches EXCLUSIVELY grey literature using cluster keyword sets.

Grey literature = reports, policy briefs, working papers, theses, parliamentary
documents, think-tank publications — NOT peer-reviewed journal articles.

Sources (all filtered to non-journal document types):
  1. OpenAlex       — type:report|preprint (journal-articles excluded)
  2. BASE           — Bielefeld Academic Search Engine, filtered to reports/working papers
  3. CORE           — filtered to reports and working papers
  4. OpenAire       — European open research infrastructure, type=report
  5. DART-Europe    — European doctoral theses
  6. EP Think Tank  — European Parliament research reports
  7. arXiv          — preprints (cs.CY, cs.SI, econ categories)

Output: grey_literature/ folder + grey_literature.xlsx (one sheet, deduplicated,
        Scopus duplicates excluded)

Requirements: pip install requests openpyxl
VPN: run with university VPN active for maximum PDF access.
"""

import os
import re
import time
import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from urllib.parse import quote_plus
import xml.etree.ElementTree as ET

# ─── CONFIGURATION ─────────────────────────────────────────────────────────────
CORE_API_KEY    = "pcjCa9P8V74fHKlGNT503RqD6nLSzwZg"
UNPAYWALL_EMAIL = "sergiopedro@ua.pt"
SCOPUS_EXCEL    = "Nets4Dem_ScopusReview_Final_GA_12 June.xlsx"
SCOPUS_SHEETS   = ["HIGH relevance", "MEDIUM screening"]
OUTPUT_FOLDER   = "grey literature"
OUTPUT_EXCEL    = "grey_literature.xlsx"
MIN_YEAR        = 2005
MAX_RESULTS_PER_SOURCE_PER_CLUSTER = 40
DELAY_SECONDS   = 1
# ───────────────────────────────────────────────────────────────────────────────

# ─── CLUSTER KEYWORD SETS ──────────────────────────────────────────────────────
CLUSTERS = {
    "Cluster 1 – Network types": {
        "group1": [
            "network of networks", "multi-actor network", "civil society hub",
            "civic ecosystem", "civic alliance", "civic platform",
            "hub organisation", "umbrella organisation", "networked organisation",
            "network formalisation",
        ],
        "group2": [
            "civil society", "democratic governance", "democratic innovation",
            "participatory democracy", "deliberation",
        ],
    },
    "Cluster 2 – Democratic intensification": {
        "group1": [
            "democratic intensification", "democratic densification",
            "shrinking civic space", "democratic backsliding",
            "democratic erosion", "de-democratization", "autocratization",
            "illiberal turn", "democratic resilience", "democratic breakdown",
        ],
        "group2": [
            "civil society", "network", "alliance", "civic organisation",
        ],
    },
    "Cluster 3 – Network resilience": {
        "group1": [
            "network resilience", "network formation", "network transformation",
            "network governance", "collaborative governance", "brokerage",
            "bridging organisations", "structural holes", "knowledge brokering",
            "co-opetition", "coopetition", "collaborative competition",
        ],
        "group2": [
            "civil society", "democratic", "civic", "participation",
        ],
    },
    "Cluster 4 – Internal governance": {
        "group1": [
            "network membership", "civil society funding", "legal form",
            "network institutionalisation", "internal governance",
            "variable geometry", "network sustainability", "informal network",
            "network formalization", "funding dependency", "governance structure",
        ],
        "group2": [
            "civil society", "democratic", "civic",
        ],
    },
    "Cluster 5 – Transnational networks Europe": {
        "group1": [
            "transnational civil society", "European civil society",
            "cross-border network", "multi-level governance",
            "pan-European network", "EU-level civil society",
            "Open Government Partnership", "civic coalition",
        ],
        "group2": [
            "Europe", "European", "EU",
        ],
    },
    "Cluster 6 – Grey literature signals": {
        "group1": [
            "civic space monitoring", "civil society index",
            "everyday democracy index", "network impact evaluation",
            "CIVICUS monitor", "network self-assessment",
            "democratic innovation report", "NGO sustainability index",
            "civic participation report", "civil society shrinking",
        ],
        "group2": [],
    },
}
# ───────────────────────────────────────────────────────────────────────────────

BASE_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept": "application/json",
    "Accept-Language": "en-US,en;q=0.5",
}

SESSION = requests.Session()
SESSION.headers.update(BASE_HEADERS)


# ── Helpers ────────────────────────────────────────────────────────────────────

def normalise_doi(raw: str) -> str | None:
    if not raw:
        return None
    doi = str(raw).strip().lower()
    for prefix in ("https://doi.org/", "http://doi.org/", "doi.org/", "doi:", "doi "):
        if doi.startswith(prefix):
            doi = doi[len(prefix):]
    doi = doi.strip()
    return doi if doi and doi != "nan" else None


def normalise_title(title: str) -> str:
    return re.sub(r'\W+', ' ', str(title).lower()).strip()


def sanitize_filename(s: str) -> str:
    return re.sub(r'[\\/:*?"<>|]', "_", s)[:180]


def cluster_query_string(cluster: dict, max_terms: int = 5) -> str:
    parts = []
    for grp_terms in cluster.values():
        if grp_terms:
            sample = grp_terms[:max_terms]
            parts.append(" OR ".join(f'"{t}"' for t in sample))
    return " AND ".join(f"({p})" for p in parts)


def try_download_pdf(url: str, referer: str = "") -> bytes | None:
    if not url:
        return None
    try:
        hdrs = {**BASE_HEADERS, "Accept": "application/pdf,*/*", "Referer": referer}
        r = SESSION.get(url, headers=hdrs, timeout=40, allow_redirects=True)
        if r.status_code == 200 and "pdf" in r.headers.get("Content-Type", "").lower():
            if len(r.content) > 5_000:
                return r.content
    except Exception:
        pass
    return None


def save_pdf(pdf_bytes: bytes, doi: str, title: str) -> str:
    os.makedirs(OUTPUT_FOLDER, exist_ok=True)
    key = sanitize_filename(doi if doi else normalise_title(title)[:80])
    path = os.path.join(OUTPUT_FOLDER, key + ".pdf")
    with open(path, "wb") as f:
        f.write(pdf_bytes)
    return path


# ── Load existing Scopus DOIs ──────────────────────────────────────────────────

def load_scopus_dois(excel_file: str) -> set:
    existing = set()
    if not os.path.exists(excel_file):
        return existing
    try:
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        for sheet in SCOPUS_SHEETS:
            if sheet not in wb.sheetnames:
                continue
            ws = wb[sheet]
            headers = [c.value for c in ws[1]]
            doi_idx = next((i for i, h in enumerate(headers)
                            if h and str(h).strip().upper() == "DOI"), None)
            if doi_idx is None:
                continue
            for row in ws.iter_rows(min_row=2, values_only=True):
                d = normalise_doi(row[doi_idx])
                if d:
                    existing.add(d)
    except Exception as e:
        print(f"  [WARN] Could not read Scopus Excel: {e}")
    return existing


# ── Search functions — grey literature only ────────────────────────────────────

def search_openalex_grey(query: str, cluster_name: str) -> list[dict]:
    """
    OpenAlex filtered to report, preprint, book — journal articles explicitly excluded.
    https://docs.openalex.org/api-entities/works/filter-works#type
    """
    results = []
    url = "https://api.openalex.org/works"
    # Include only non-journal types
    type_filter = "report|preprint|book|book-chapter|dissertation|dataset"
    params = {
        "search": query,
        "filter": f"publication_year:>{MIN_YEAR - 1},type:{type_filter}",
        "per-page": MAX_RESULTS_PER_SOURCE_PER_CLUSTER,
        "select": "doi,title,publication_year,authorships,primary_location,open_access,type",
    }
    try:
        r = SESSION.get(url, params=params, timeout=20)
        if r.status_code != 200:
            return results
        for item in r.json().get("results") or []:
            doi = normalise_doi(item.get("doi") or "")
            oa = item.get("open_access") or {}
            loc = item.get("primary_location") or {}
            authors = "; ".join(
                a.get("author", {}).get("display_name", "") or ""
                for a in (item.get("authorships") or [])[:5]
            )
            results.append({
                "doi": doi or "",
                "title": item.get("title") or "",
                "year": item.get("publication_year") or "",
                "authors": authors,
                "doc_type": item.get("type") or "unknown",
                "source_db": "OpenAlex (grey)",
                "cluster": cluster_name,
                "abstract": "",
                "landing_url": loc.get("landing_page_url") or "",
                "pdf_url": oa.get("oa_url") or loc.get("pdf_url") or "",
            })
    except Exception as e:
        print(f"    [OpenAlex error] {e}")
    return results


def search_base_grey(query: str, cluster_name: str) -> list[dict]:
    """
    BASE — Bielefeld Academic Search Engine.
    Filtered to doctypes: 4=report, 15=working paper, 6=thesis, 14=conference paper.
    https://www.base-search.net/about/en/faq.php
    """
    results = []
    url = "https://api.base-search.net/cgi-bin/BaseHttpSearchInterface.fcgi"
    # dcdoctype codes: 1=article, 4=report, 5=book, 6=thesis, 14=conference, 15=working paper
    params = {
        "func": "PerformSearch",
        "query": f"dcterms.description:({query}) AND (dcdoctype:4 OR dcdoctype:6 OR dcdoctype:15 OR dcdoctype:14)",
        "hits": MAX_RESULTS_PER_SOURCE_PER_CLUSTER,
        "offset": 0,
        "format": "json",
        "boost": "oa",
    }
    try:
        r = SESSION.get(url, params=params, timeout=20)
        if r.status_code != 200:
            return results
        docs = r.json().get("response", {}).get("docs") or []
        doctype_labels = {
            "1": "article", "4": "report", "5": "book",
            "6": "thesis", "14": "conference paper", "15": "working paper",
        }
        for item in docs:
            year_raw = str(item.get("dcyear") or "")[:4]
            year = int(year_raw) if year_raw.isdigit() else 0
            if year and year < MIN_YEAR:
                continue
            doi = normalise_doi(item.get("dcdoi") or "")
            links = item.get("dclink") or []
            pdf_url = next((l for l in links if ".pdf" in l.lower()), "")
            landing = links[0] if links else ""
            authors_raw = item.get("dcauthor") or []
            authors = "; ".join(authors_raw[:5]) if isinstance(authors_raw, list) else str(authors_raw)
            dtype_code = str(item.get("dcdoctype") or "")
            dtype = doctype_labels.get(dtype_code, f"type:{dtype_code}")
            title_raw = item.get("dctitle") or ""
            title = (title_raw[0] if isinstance(title_raw, list) else str(title_raw))
            results.append({
                "doi": doi or "",
                "title": title,
                "year": year,
                "authors": authors,
                "doc_type": dtype,
                "source_db": "BASE",
                "cluster": cluster_name,
                "abstract": str(item.get("dcdescription") or "")[:400],
                "landing_url": landing,
                "pdf_url": pdf_url,
            })
    except Exception as e:
        print(f"    [BASE error] {e}")
    return results


def search_core_grey(query: str, cluster_name: str) -> list[dict]:
    """
    CORE API — working papers, reports, theses. Filters out journal articles
    by checking document type field where available.
    """
    results = []
    url = "https://api.core.ac.uk/v3/search/works"
    params = {
        "q": f"{query} AND (documentType:report OR documentType:thesis OR documentType:workingPaper OR documentType:conferenceProceedings)",
        "limit": MAX_RESULTS_PER_SOURCE_PER_CLUSTER,
        "offset": 0,
        "yearFrom": MIN_YEAR,
    }
    hdrs = {**BASE_HEADERS, "Authorization": f"Bearer {CORE_API_KEY}"}
    try:
        r = SESSION.get(url, headers=hdrs, params=params, timeout=20)
        if r.status_code != 200:
            # fallback: search without type filter, skip if doctype is journal
            params["q"] = query
            r = SESSION.get(url, headers=hdrs, params=params, timeout=20)
            if r.status_code != 200:
                return results
        for item in r.json().get("results") or []:
            year = item.get("yearPublished") or 0
            if year and int(year) < MIN_YEAR:
                continue
            dtype = (item.get("documentType") or "").lower()
            # Skip if explicitly a journal article
            if dtype in ("journal article", "journalarticle", "article"):
                continue
            doi = normalise_doi(item.get("doi") or "")
            pdf_url = item.get("downloadUrl") or ""
            landing = (item.get("sourceFulltextUrls") or [None])[0] or ""
            authors = "; ".join(
                (a.get("name") or "") for a in (item.get("authors") or [])
            )
            results.append({
                "doi": doi or "",
                "title": item.get("title") or "",
                "year": year,
                "authors": authors,
                "doc_type": dtype or "working paper / report",
                "source_db": "CORE",
                "cluster": cluster_name,
                "abstract": (item.get("abstract") or "")[:400],
                "landing_url": landing,
                "pdf_url": pdf_url,
            })
    except Exception as e:
        print(f"    [CORE error] {e}")
    return results


def search_openaire_grey(query: str, cluster_name: str) -> list[dict]:
    """
    OpenAire — European open research infrastructure.
    Restricted to type=report (excludes journal articles).
    https://graph.openaire.eu/develop/api.html
    """
    results = []
    url = "https://api.openaire.eu/search/publications"
    params = {
        "keywords": query,
        "type": "report",
        "size": MAX_RESULTS_PER_SOURCE_PER_CLUSTER,
        "format": "json",
        "fromDateAccepted": f"{MIN_YEAR}-01-01",
    }
    try:
        r = SESSION.get(url, params=params, timeout=20,
                        headers={**BASE_HEADERS, "Accept": "application/json"})
        if r.status_code != 200:
            return results
        response = r.json().get("response") or {}
        results_list = response.get("results") or {}
        items = results_list.get("result") or []
        if isinstance(items, dict):
            items = [items]
        for item in items:
            metadata = item.get("metadata", {}).get("oaf:entity", {}).get("oaf:result", {})
            title_raw = metadata.get("title") or {}
            title = title_raw.get("$", "") if isinstance(title_raw, dict) else str(title_raw)
            year = str(metadata.get("dateofacceptance") or "")[:4]
            doi = ""
            for pid in (metadata.get("pid") or []):
                if isinstance(pid, dict) and pid.get("@classid") == "doi":
                    doi = normalise_doi(pid.get("$", "")) or ""
                    break
            pdf_url = ""
            for inst in (metadata.get("instance") or []):
                if isinstance(inst, dict):
                    for url_item in (inst.get("url") or []):
                        if isinstance(url_item, dict) and ".pdf" in url_item.get("$", "").lower():
                            pdf_url = url_item.get("$", "")
                            break
            results.append({
                "doi": doi or "",
                "title": title,
                "year": year,
                "authors": "",
                "doc_type": "report",
                "source_db": "OpenAire",
                "cluster": cluster_name,
                "abstract": "",
                "landing_url": "",
                "pdf_url": pdf_url,
            })
    except Exception as e:
        print(f"    [OpenAire error] {e}")
    return results


def search_dart_europe(query: str, cluster_name: str) -> list[dict]:
    """
    DART-Europe — European doctoral theses repository.
    OAI-PMH endpoint, simple keyword search.
    http://www.dart-europe.org/basic-search.php
    """
    results = []
    url = "https://www.dart-europe.org/basic-search.php"
    params = {
        "query": query,
        "format": "json",
        "rows": MAX_RESULTS_PER_SOURCE_PER_CLUSTER,
    }
    try:
        r = SESSION.get(url, params=params, timeout=20)
        if r.status_code != 200:
            return results
        data = r.json()
        for item in (data.get("docs") or data.get("results") or []):
            year_raw = str(item.get("year") or item.get("date") or "")[:4]
            year = int(year_raw) if year_raw.isdigit() else 0
            if year and year < MIN_YEAR:
                continue
            results.append({
                "doi": normalise_doi(item.get("doi") or "") or "",
                "title": item.get("title") or "",
                "year": year,
                "authors": item.get("author") or "",
                "doc_type": "doctoral thesis",
                "source_db": "DART-Europe",
                "cluster": cluster_name,
                "abstract": (item.get("abstract") or "")[:400],
                "landing_url": item.get("url") or item.get("link") or "",
                "pdf_url": item.get("pdf") or "",
            })
    except Exception as e:
        print(f"    [DART-Europe error] {e}")
    return results


def search_ep_think_tank(query: str, cluster_name: str) -> list[dict]:
    """
    European Parliament Think Tank — parliamentary research reports and briefings.
    https://www.europarl.europa.eu/thinktank/
    """
    results = []
    url = "https://www.europarl.europa.eu/thinktank/en/search.html"
    params = {
        "query": query,
        "format": "json",
        "rows": MAX_RESULTS_PER_SOURCE_PER_CLUSTER,
        "start": 0,
    }
    try:
        r = SESSION.get(url, params=params, timeout=20,
                        headers={**BASE_HEADERS, "Accept": "application/json, text/javascript"})
        if r.status_code != 200:
            return results
        data = r.json()
        for item in (data.get("docs") or data.get("results") or data.get("items") or []):
            year_raw = str(item.get("date") or item.get("year") or "")[:4]
            year = int(year_raw) if year_raw.isdigit() else 0
            if year and year < MIN_YEAR:
                continue
            pdf_url = item.get("pdfUrl") or item.get("pdf_url") or ""
            landing = item.get("url") or item.get("link") or ""
            results.append({
                "doi": "",
                "title": item.get("title") or "",
                "year": year,
                "authors": item.get("authors") or item.get("author") or "European Parliament",
                "doc_type": item.get("type") or "EP research report",
                "source_db": "EP Think Tank",
                "cluster": cluster_name,
                "abstract": (item.get("summary") or item.get("abstract") or "")[:400],
                "landing_url": landing,
                "pdf_url": pdf_url,
            })
    except Exception as e:
        print(f"    [EP Think Tank error] {e}")
    return results


def search_arxiv_preprints(query: str, cluster_name: str) -> list[dict]:
    """
    arXiv — preprints in relevant social science / computer science categories.
    Clearly labelled as preprints (not peer-reviewed).
    """
    results = []
    cats = "cat:cs.CY OR cat:cs.SI OR cat:econ.GN OR cat:econ.PE OR cat:soc.TH"
    url = "http://export.arxiv.org/api/query"
    params = {
        "search_query": f"all:({query}) AND ({cats})",
        "max_results": MAX_RESULTS_PER_SOURCE_PER_CLUSTER,
        "sortBy": "relevance",
    }
    try:
        r = SESSION.get(url, params=params, timeout=20,
                        headers={**BASE_HEADERS, "Accept": "application/xml"})
        if r.status_code != 200:
            return results
        ns = {"a": "http://www.w3.org/2005/Atom"}
        root = ET.fromstring(r.text)
        for entry in root.findall("a:entry", ns):
            year_raw = (entry.findtext("a:published", "", ns) or "")[:4]
            year = int(year_raw) if year_raw.isdigit() else 0
            if year and year < MIN_YEAR:
                continue
            arxiv_id = entry.findtext("a:id", "", ns).split("/abs/")[-1]
            authors = "; ".join(
                a.findtext("a:name", "", ns)
                for a in entry.findall("a:author", ns)[:5]
            )
            results.append({
                "doi": "",
                "title": entry.findtext("a:title", "", ns).replace("\n", " ").strip(),
                "year": year,
                "authors": authors,
                "doc_type": "preprint",
                "source_db": "arXiv (preprint)",
                "cluster": cluster_name,
                "abstract": entry.findtext("a:summary", "", ns).replace("\n", " ").strip()[:400],
                "landing_url": f"https://arxiv.org/abs/{arxiv_id}",
                "pdf_url": f"https://arxiv.org/pdf/{arxiv_id}",
            })
    except Exception as e:
        print(f"    [arXiv error] {e}")
    return results


SEARCH_FUNCTIONS = [
    ("OpenAlex (grey)",  search_openalex_grey),
    ("BASE",             search_base_grey),
    ("CORE",             search_core_grey),
    ("OpenAire",         search_openaire_grey),
    ("DART-Europe",      search_dart_europe),
    ("EP Think Tank",    search_ep_think_tank),
    ("arXiv (preprint)", search_arxiv_preprints),
]


# ── Deduplication ──────────────────────────────────────────────────────────────

def deduplicate(records: list[dict], scopus_dois: set) -> list[dict]:
    seen_dois = set(scopus_dois)
    seen_titles = set()
    unique = []
    for rec in records:
        doi = rec.get("doi", "").strip()
        title_key = normalise_title(rec.get("title", ""))
        if not title_key or len(title_key) < 5:
            continue
        if doi and doi in seen_dois:
            continue
        if title_key and title_key in seen_titles:
            continue
        if doi:
            seen_dois.add(doi)
        if title_key:
            seen_titles.add(title_key)
        unique.append(rec)
    return unique


# ── Excel output ───────────────────────────────────────────────────────────────

EXCEL_COLUMNS = [
    "title", "authors", "year", "doc_type", "doi", "cluster",
    "source_db", "abstract", "landing_url", "pdf_url", "pdf_downloaded",
]

def write_excel(records: list[dict]):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "grey literature"

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    headers_display = [
        "Title", "Authors", "Year", "Document Type", "DOI", "Cluster",
        "Source DB", "Abstract", "Landing URL", "PDF URL", "PDF Downloaded",
    ]
    ws.append(headers_display)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for rec in records:
        ws.append([rec.get(c, "") for c in EXCEL_COLUMNS])

    col_widths = [60, 35, 6, 20, 35, 30, 20, 70, 60, 60, 15]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    wb.save(OUTPUT_EXCEL)
    print(f"\nExcel saved: {os.path.abspath(OUTPUT_EXCEL)}")


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    print(f"Loading existing Scopus DOIs from '{SCOPUS_EXCEL}' ...")
    scopus_dois = load_scopus_dois(SCOPUS_EXCEL)
    print(f"  {len(scopus_dois)} DOIs loaded (will be excluded as already in Scopus).\n")

    os.makedirs(OUTPUT_FOLDER, exist_ok=True)

    all_records = []

    for cluster_name, cluster in CLUSTERS.items():
        print(f"── {cluster_name} ──────────────────────────────────────────")
        query = cluster_query_string(cluster)
        print(f"  Query: {query[:120]}...")

        for src_name, fn in SEARCH_FUNCTIONS:
            print(f"  Searching {src_name} ...", end=" ", flush=True)
            results = fn(query, cluster_name)
            print(f"{len(results)} results")
            all_records.extend(results)
            time.sleep(DELAY_SECONDS)

    print(f"\nTotal raw results: {len(all_records)}")
    unique = deduplicate(all_records, scopus_dois)
    print(f"After deduplication (Scopus excluded): {len(unique)}")

    # Summary by doc_type
    from collections import Counter
    type_counts = Counter(r.get("doc_type", "unknown") for r in unique)
    print("\nDocument types found:")
    for dtype, count in type_counts.most_common():
        print(f"  {dtype}: {count}")

    # Download PDFs
    print(f"\nAttempting PDF downloads into '{OUTPUT_FOLDER}/' ...")
    for i, rec in enumerate(unique, start=1):
        pdf_url = rec.get("pdf_url") or rec.get("landing_url") or ""
        doi = rec.get("doi", "")
        title = rec.get("title", f"record_{i}")
        filename_key = sanitize_filename(doi if doi else normalise_title(title)[:80])
        path = os.path.join(OUTPUT_FOLDER, filename_key + ".pdf")

        if os.path.exists(path):
            rec["pdf_downloaded"] = "YES (existing)"
            continue

        if not pdf_url:
            rec["pdf_downloaded"] = "NO – no URL"
            continue

        print(f"  [{i}/{len(unique)}] {title[:60]} ...", end=" ", flush=True)
        pdf = try_download_pdf(pdf_url, referer=f"https://doi.org/{doi}" if doi else pdf_url)

        # Unpaywall fallback if DOI available
        if not pdf and doi:
            try:
                r = SESSION.get(
                    f"https://api.unpaywall.org/v2/{doi}?email={UNPAYWALL_EMAIL}",
                    timeout=10
                )
                if r.status_code == 200:
                    for loc in (r.json().get("oa_locations") or []):
                        u = loc.get("url_for_pdf") or loc.get("url")
                        if u:
                            pdf = try_download_pdf(u, referer=f"https://doi.org/{doi}")
                            if pdf:
                                break
            except Exception:
                pass

        if pdf:
            save_pdf(pdf, doi if doi else f"record_{i}", title)
            rec["pdf_downloaded"] = "YES"
            print("OK")
        else:
            rec["pdf_downloaded"] = "NO"
            print("FAIL")

        time.sleep(DELAY_SECONDS)

    write_excel(unique)

    downloaded = sum(1 for r in unique if r.get("pdf_downloaded", "").startswith("YES"))
    print(f"\n{'='*60}")
    print(f"Total unique grey literature results: {len(unique)}")
    print(f"PDFs downloaded: {downloaded}")
    print(f"Output folder:   {os.path.abspath(OUTPUT_FOLDER)}")
    print(f"Output Excel:    {os.path.abspath(OUTPUT_EXCEL)}")
    print(f"\nNOTE: For PDFs not downloaded automatically, use the 'Landing URL'")
    print(f"column in the Excel to access them manually via your university VPN.")


if __name__ == "__main__":
    main()
