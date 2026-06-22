"""
Compares the Excel highlighted DOIs against downloaded PDFs
and writes missing ones to failed_downloads.csv
"""

import os
import re
import csv
import openpyxl

# ─── CONFIGURATION (must match download_pdfs.py) ──────────────────────────────
EXCEL_FILE     = "Nets4Dem_ScopusReview_Final_GA_12 June.xlsx"
DOI_COLUMN     = "DOI"
YELLOW_COLOR   = "FFFFFF00"
GREEN_COLORS   = {"FFC6EFCE", "FF92D050"}
SHEETS_TO_SCAN = ["HIGH relevance", "MEDIUM screening"]
FUTURE_FOLDER  = "Future readings"
OUTPUT_CSV     = "failed_downloads.csv"
# ───────────────────────────────────────────────────────────────────────────────


def clean_doi(raw) -> str | None:
    if raw is None:
        return None
    doi = str(raw).strip()
    for prefix in ("https://doi.org/", "http://doi.org/", "doi.org/", "DOI:", "doi:"):
        if doi.lower().startswith(prefix.lower()):
            doi = doi[len(prefix):]
    doi = doi.strip()
    return doi if doi and doi.lower() != "nan" else None


def sanitize_filename(doi: str) -> str:
    return re.sub(r'[\\/:*?"<>|]', "_", doi)


def row_color(ws_row) -> str | None:
    for cell in ws_row:
        fill = cell.fill
        if fill and fill.fgColor and fill.fgColor.type == "rgb":
            rgb = fill.fgColor.rgb
            if rgb and rgb not in ("00000000", "FF000000"):
                return rgb
    return None


def main():
    if not os.path.exists(EXCEL_FILE):
        print(f"[ERROR] '{EXCEL_FILE}' not found.")
        return

    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    failed = []
    total = 0

    for sheet_name in SHEETS_TO_SCAN:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        header = [cell.value for cell in ws[1]]
        try:
            doi_idx = header.index(DOI_COLUMN)
        except ValueError:
            print(f"[WARN] Column '{DOI_COLUMN}' not in '{sheet_name}'")
            continue

        for row in ws.iter_rows(min_row=2):
            color = row_color(row)
            doi = clean_doi(row[doi_idx].value)
            if not doi:
                continue

            if color == YELLOW_COLOR:
                folder = sheet_name
            elif color in GREEN_COLORS:
                folder = FUTURE_FOLDER
            else:
                continue

            total += 1
            path = os.path.join(folder, sanitize_filename(doi) + ".pdf")
            if not os.path.exists(path):
                failed.append({
                    "doi": doi,
                    "folder": folder,
                    "doi_url": f"https://doi.org/{doi}"
                })

    with open(OUTPUT_CSV, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=["doi", "folder", "doi_url"])
        writer.writeheader()
        writer.writerows(failed)

    print(f"Total expected: {total}")
    print(f"Missing PDFs:   {len(failed)}")
    print(f"Saved to:       {os.path.abspath(OUTPUT_CSV)}")


if __name__ == "__main__":
    main()
