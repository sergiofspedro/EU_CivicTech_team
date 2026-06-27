"""
build_batch_excel.py
Creates two Excel files with one sheet per NotebookLM batch:

  batch_overview.xlsx       — Batches 1–5 (Scopus articles)
  batch_overview_grey.xlsx  — Batches 6–7 (grey literature)

Metadata for batches 1–5 comes from the Scopus Excel.
Metadata for batches 6–7 comes from grey_literature.xlsx.
PDF filenames in all folders are the sanitized DOI (or title key).

Usage:
  python build_batch_excel.py
"""

import os
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ─── CONFIGURATION ─────────────────────────────────────────────────────────────
SCOPUS_EXCEL      = "Nets4Dem_ScopusReview_Final_GA_12 June.xlsx"
SCOPUS_SHEETS     = ["HIGH relevance", "MEDIUM screening"]
GREY_EXCEL        = "grey_literature.xlsx"
GREY_SHEET        = "grey literature"

OUTPUT_SCOPUS     = "batch_overview.xlsx"
OUTPUT_GREY       = "batch_overview_grey.xlsx"

# Batches 1–5: Scopus articles
SCOPUS_BATCHES = [
    ("Batch 1", r"C:\Users\Administrator\Downloads\Livro\HIGH relevance\1"),
    ("Batch 2", r"C:\Users\Administrator\Downloads\Livro\HIGH relevance\2"),
    ("Batch 3", r"C:\Users\Administrator\Downloads\Livro\HIGH relevance\3"),
    ("Batch 4", r"C:\Users\Administrator\Downloads\Livro\MEDIUM screening\4"),
    ("Batch 5", r"C:\Users\Administrator\Downloads\Livro\MEDIUM screening\5"),
]

# Batches 6–7: grey literature
GREY_BATCHES = [
    ("Batch 6", r"C:\Users\Administrator\Downloads\Livro\grey literature\6"),
    ("Batch 7", r"C:\Users\Administrator\Downloads\Livro\grey literature\7"),
]
# ───────────────────────────────────────────────────────────────────────────────


def sanitize_filename(s: str) -> str:
    return re.sub(r'[\\/:*?"<>|]', "_", s)


def get_pdf_keys(folder_path: str) -> list[str]:
    """List all PDF filenames (without extension) in a folder."""
    if not os.path.isdir(folder_path):
        print(f"  [WARN] Folder not found: {folder_path}")
        return []
    return sorted(f[:-4] for f in os.listdir(folder_path) if f.lower().endswith(".pdf"))


def write_sheet(ws, headers: list, rows: list):
    """Write headers + rows with styling."""
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)

    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for row in rows:
        ws.append(row)

    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = len(str(headers[col_idx - 1]))
        for row_cells in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row_cells:
                if cell.value:
                    max_len = max(max_len, min(len(str(cell.value)), 60))
        ws.column_dimensions[col_letter].width = max_len + 2

    ws.freeze_panes = "A2"


# ── Scopus lookup ──────────────────────────────────────────────────────────────

def build_scopus_lookup(excel_file: str) -> tuple[dict, list]:
    """Returns (lookup_by_sanitized_doi, all_column_names)."""
    lookup = {}
    all_cols = []
    seen_cols = set()

    wb = openpyxl.load_workbook(excel_file, data_only=True)
    for sheet_name in SCOPUS_SHEETS:
        if sheet_name not in wb.sheetnames:
            print(f"  [WARN] Sheet '{sheet_name}' not found in {excel_file}")
            continue
        ws = wb[sheet_name]
        headers = [cell.value for cell in ws[1]]

        doi_idx = next((i for i, h in enumerate(headers)
                        if h and str(h).strip().upper() == "DOI"), None)
        if doi_idx is None:
            print(f"  [WARN] No DOI column in '{sheet_name}'")
            continue

        for h in headers:
            if h and str(h) not in seen_cols:
                all_cols.append(str(h))
                seen_cols.add(str(h))

        for row in ws.iter_rows(min_row=2, values_only=True):
            raw = row[doi_idx]
            if not raw:
                continue
            doi = str(raw).strip()
            for prefix in ("https://doi.org/", "http://doi.org/", "doi.org/", "DOI:", "doi:"):
                if doi.lower().startswith(prefix.lower()):
                    doi = doi[len(prefix):]
            doi = doi.strip()
            if not doi or doi.lower() == "nan":
                continue
            key = sanitize_filename(doi)
            record = {"_source_sheet": sheet_name}
            for col_name, value in zip(headers, row):
                if col_name:
                    record[str(col_name)] = value
            lookup[key] = record

    print(f"  Loaded {len(lookup)} Scopus articles.")
    return lookup, all_cols


# ── Grey literature lookup ─────────────────────────────────────────────────────

def build_grey_lookup(excel_file: str) -> tuple[dict, list]:
    """
    Reads grey_literature.xlsx and builds a lookup by sanitized filename key.
    The filename key is the sanitized DOI if available, otherwise sanitized title.
    """
    lookup = {}
    all_cols = []

    if not os.path.exists(excel_file):
        print(f"  [WARN] {excel_file} not found — grey batches will show unmatched.")
        return lookup, all_cols

    wb = openpyxl.load_workbook(excel_file, data_only=True)
    if GREY_SHEET not in wb.sheetnames:
        print(f"  [WARN] Sheet '{GREY_SHEET}' not found in {excel_file}")
        return lookup, all_cols

    ws = wb[GREY_SHEET]
    headers = [str(cell.value) for cell in ws[1] if cell.value]
    all_cols = headers

    # Find DOI and Title column indices
    doi_idx   = next((i for i, h in enumerate(headers) if h.upper() == "DOI"), None)
    title_idx = next((i for i, h in enumerate(headers) if h.upper() == "TITLE"), None)

    for row in ws.iter_rows(min_row=2, values_only=True):
        record = {h: row[i] for i, h in enumerate(headers) if i < len(row)}

        # Build the key: sanitized DOI if available, else sanitized title (truncated)
        doi_raw = str(row[doi_idx] or "").strip() if doi_idx is not None else ""
        title_raw = str(row[title_idx] or "").strip() if title_idx is not None else ""

        if doi_raw and doi_raw.lower() not in ("", "nan", "none"):
            key = sanitize_filename(doi_raw)[:180]
        elif title_raw and title_raw.lower() not in ("", "nan", "none"):
            clean = re.sub(r'\W+', ' ', title_raw.lower()).strip()
            key = sanitize_filename(clean)[:80]
        else:
            continue

        lookup[key] = record

    print(f"  Loaded {len(lookup)} grey literature records.")
    return lookup, all_cols


# ── Build one Excel file ───────────────────────────────────────────────────────

def build_excel(batches: list, lookup: dict, all_cols: list,
                output_file: str, source_col_label: str):

    output_headers = ([source_col_label] + all_cols
                      + ["PDF filename", "Matched"])

    out_wb = openpyxl.Workbook()
    out_wb.remove(out_wb.active)

    summary_rows = []

    for sheet_name, folder_path in batches:
        print(f"\n  Batch '{sheet_name}' ← {folder_path}")
        keys = get_pdf_keys(folder_path)
        print(f"    {len(keys)} PDFs found")

        rows = []
        matched = unmatched = 0
        for key in keys:
            record = lookup.get(key)
            if record:
                row = [record.get("_source_sheet", record.get(source_col_label, ""))]
                for col in all_cols:
                    row.append(record.get(col, ""))
                row += [key + ".pdf", "YES"]
                matched += 1
            else:
                row = [""] * (len(all_cols) + 1) + [key + ".pdf", "NO – not in metadata"]
                unmatched += 1
            rows.append(row)

        ws = out_wb.create_sheet(title=sheet_name[:31])
        write_sheet(ws, output_headers, rows)
        print(f"    Matched: {matched}  |  Unmatched: {unmatched}")
        summary_rows.append((sheet_name, folder_path, len(keys), matched, unmatched))

    # Summary sheet
    ws_sum = out_wb.create_sheet(title="Summary", index=0)
    write_sheet(ws_sum,
                ["Batch", "Folder", "Total PDFs", "Matched", "Unmatched"],
                summary_rows)

    out_wb.save(output_file)
    print(f"\n  Saved → {os.path.abspath(output_file)}")
    for name, _, total, m, u in summary_rows:
        print(f"    {name}: {total} PDFs | {m} matched | {u} unmatched")


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    # ── Scopus batches 1–5 ───────────────────────────────────────────────────
    print(f"\n{'='*60}")
    print(f"SCOPUS BATCHES (1–5) → {OUTPUT_SCOPUS}")
    print(f"{'='*60}")

    if not os.path.exists(SCOPUS_EXCEL):
        print(f"[ERROR] {SCOPUS_EXCEL} not found.")
    else:
        scopus_lookup, scopus_cols = build_scopus_lookup(SCOPUS_EXCEL)
        build_excel(SCOPUS_BATCHES, scopus_lookup, scopus_cols,
                    OUTPUT_SCOPUS, "Source sheet")

    # ── Grey literature batches 6–7 ──────────────────────────────────────────
    print(f"\n{'='*60}")
    print(f"GREY LITERATURE BATCHES (6–7) → {OUTPUT_GREY}")
    print(f"{'='*60}")

    grey_lookup, grey_cols = build_grey_lookup(GREY_EXCEL)
    build_excel(GREY_BATCHES, grey_lookup, grey_cols,
                OUTPUT_GREY, "Source DB")

    print(f"\nDone.")


if __name__ == "__main__":
    main()
