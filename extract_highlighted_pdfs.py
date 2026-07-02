"""
extract_highlighted_pdfs.py
Reads batch_overview.xlsx (with your yellow highlights across Batch 1-5 sheets)
and copies the corresponding PDFs from their batch folders into a single
destination folder.

Usage:
  python extract_highlighted_pdfs.py
"""

import os
import shutil
import openpyxl

# ─── CONFIGURATION ─────────────────────────────────────────────────────────────
REVIEWED_EXCEL   = "batch_overview.xlsx"   # the file you highlighted

# Source folders for each batch (must match where the PDFs actually live)
BATCH_FOLDERS = {
    "Batch 1": r"C:\Users\Administrator\Downloads\Livro\HIGH relevance\1",
    "Batch 2": r"C:\Users\Administrator\Downloads\Livro\HIGH relevance\2",
    "Batch 3": r"C:\Users\Administrator\Downloads\Livro\HIGH relevance\3",
    "Batch 4": r"C:\Users\Administrator\Downloads\Livro\MEDIUM screening\4",
    "Batch 5": r"C:\Users\Administrator\Downloads\Livro\MEDIUM screening\5",
}
BATCH_SHEETS = list(BATCH_FOLDERS.keys())

DEST_FOLDER = r"C:\Users\Administrator\Downloads\Livro\New batch"
# ───────────────────────────────────────────────────────────────────────────────


def row_is_highlighted(ws_row) -> str | None:
    """
    Returns a description of the fill found on this row, or None if no
    cell has a non-default background fill. Handles rgb, indexed, and
    theme-based fill colors (Excel's "highlight" button can use any of these).
    """
    for cell in ws_row:
        fill = cell.fill
        if not fill or fill.patternType is None:
            continue
        fg = fill.fgColor
        if fg is None:
            continue
        if fg.type == "rgb" and fg.rgb and fg.rgb not in ("00000000", "FF000000", None):
            return f"rgb:{fg.rgb}"
        if fg.type == "indexed" and fg.indexed not in (64, 65):  # 64/65 = default/none
            return f"indexed:{fg.indexed}"
        if fg.type == "theme" and fg.theme is not None:
            return f"theme:{fg.theme}/tint:{fg.tint}"
    return None


def main():
    if not os.path.exists(REVIEWED_EXCEL):
        print(f"[ERROR] '{REVIEWED_EXCEL}' not found.")
        return

    os.makedirs(DEST_FOLDER, exist_ok=True)

    wb = openpyxl.load_workbook(REVIEWED_EXCEL, data_only=True)

    copied = missing = 0
    missing_list = []
    colors_seen = {}

    for sheet_name in BATCH_SHEETS:
        if sheet_name not in wb.sheetnames:
            print(f"[WARN] Sheet '{sheet_name}' not found — skipping.")
            continue
        ws = wb[sheet_name]
        headers = [c.value for c in ws[1]]
        pdf_idx = headers.index("PDF filename") if "PDF filename" in headers else None
        title_idx = headers.index("Title") if "Title" in headers else None
        if pdf_idx is None:
            print(f"[WARN] No 'PDF filename' column in '{sheet_name}' — skipping.")
            continue

        src_folder = BATCH_FOLDERS.get(sheet_name)
        sheet_count = 0

        for row in ws.iter_rows(min_row=2):
            color = row_is_highlighted(row)
            if color is None:
                continue
            colors_seen[color] = colors_seen.get(color, 0) + 1

            pdf_name = row[pdf_idx].value
            if not pdf_name:
                continue
            pdf_name = str(pdf_name).strip()

            src_path = os.path.join(src_folder, pdf_name)
            dest_path = os.path.join(DEST_FOLDER, pdf_name)

            title = str(row[title_idx].value)[:60] if title_idx is not None and row[title_idx].value else pdf_name

            if not os.path.exists(src_path):
                print(f"  [MISSING] {sheet_name}: {pdf_name} — {title}")
                missing += 1
                missing_list.append((sheet_name, pdf_name, title))
                continue

            if os.path.exists(dest_path):
                print(f"  [SKIP] Already in destination: {pdf_name}")
            else:
                shutil.copy2(src_path, dest_path)
                print(f"  [OK] {sheet_name}: {pdf_name}")

            copied += 1
            sheet_count += 1

        print(f"{sheet_name}: {sheet_count} highlighted PDFs processed\n")

    print(f"{'='*60}")
    print(f"Total copied/found: {copied}")
    print(f"Missing source files: {missing}")
    print(f"Destination folder: {os.path.abspath(DEST_FOLDER)}")
    print(f"\nFill colors detected on highlighted rows: {colors_seen if colors_seen else '(none found)'}")
    if not colors_seen:
        print("  No highlighted cells were detected at all. Possible causes:")
        print("  - The highlights are in a different file than 'batch_overview.xlsx'")
        print("  - The file was re-saved by a tool that stripped formatting")
        print("  - You highlighted a column beyond what openpyxl scans (unlikely)")

    if missing_list:
        print(f"\nMissing files (source PDF not found in batch folder):")
        for sheet_name, pdf_name, title in missing_list:
            print(f"  {sheet_name}: {pdf_name} — {title}")


if __name__ == "__main__":
    main()
