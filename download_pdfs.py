"""
PDF Downloader — Nets4Dem Scopus Review
Downloads PDFs for highlighted rows in Excel via an 8-layer pipeline:
  1. Unpaywall     (all OA locations)
  2. CORE API      (200M+ OA papers)
  3. Semantic Scholar
  4. OpenAlex      (all locations)
  5. PubMed Central
  6. Europe PMC
  7. Sci-Hub       (improved URL parsing + title fallback)
  8. VPN/HTML      (doi.org with session + cookies)

Retry mode: if failed_downloads.csv exists, processes only those DOIs.
Outputs: failed_downloads.csv, download_log.csv
"""

import os
import re
import csv
import time
import requests
import openpyxl
from urllib.parse import urljoin, urlparse, quote_plus

# ─── CONFIGURATION ─────────────────────────────────────────────────────────────
EXCEL_FILE      = "Nets4Dem_ScopusReview_Final_GA_12 June.xlsx"
DOI_COLUMN      = "DOI"
TITLE_COLUMN    = "Title"
UNPAYWALL_EMAIL = "sergiopedro@ua.pt"
CORE_API_KEY    = "pcjCa9P8V74fHKlGNT503RqD6nLSzwZg"
DELAY_SECONDS   = 2
RETRY_TIMES     = 2       # retries per layer on timeout/connection error
RETRY_DELAY     = 3       # seconds between retries (doubles each time)

YELLOW_COLOR    = "FFFFFF00"
GREEN_COLORS    = {"FFC6EFCE", "FF92D050"}
SHEETS_TO_SCAN  = ["HIGH relevance", "MEDIUM screening"]
FUTURE_FOLDER   = "Future readings"

SCIHUB_MIRRORS  = [
    "https://sci-hub.se",
    "https://sci-hub.st",
    "https://sci-hub.ru",
]
# ───────────────────────────────────────────────────────────────────────────────

BASE_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.5",
    "Accept-Encoding": "gzip, deflate, br",
    "Connection": "keep-alive",
    "Upgrade-Insecure-Requests": "1",
}

PDF_HEADERS = {
    **BASE_HEADERS,
    "Accept": "application/pdf,application/octet-stream,*/*;q=0.8",
}


# ── Session (shared across all requests for cookie persistence) ────────────────
SESSION = requests.Session()
SESSION.headers.update(BASE_HEADERS)


# ── Helpers ────────────────────────────────────────────────────────────────────

def clean_doi(raw) -> str | None:
    if raw is None:
        return None
    doi = str(raw).strip()
    for prefix in ("https://doi.org/", "http://doi.org/", "doi.org/", "DOI:", "doi:"):
        if doi.lower().startswith(prefix.lower()):
            doi = doi[len(prefix):]
    doi = doi.strip()
    return doi if doi and doi.lower() != "nan" else None


def sanitize_filename(doi: str) -> str:
    return re.sub(r'[\\/:*?"<>|]', "_", doi)


def row_color(ws_row) -> str | None:
    for cell in ws_row:
        fill = cell.fill
        if fill and fill.fgColor and fill.fgColor.type == "rgb":
            rgb = fill.fgColor.rgb
            if rgb and rgb not in ("00000000", "FF000000"):
                return rgb
    return None


def make_absolute(url: str, base: str) -> str | None:
    if not url:
        return None
    if url.startswith("//"):
        scheme = urlparse(base).scheme or "https"
        url = f"{scheme}:{url}"
    full = urljoin(base, url)
    return full if urlparse(full).scheme in ("http", "https") else None


def fetch(url: str, timeout: int = 20, headers: dict = None,
          referer: str = None) -> tuple[requests.Response | None, str]:
    """Fetch with retry on timeout/connection error. Returns (response, error)."""
    hdrs = dict(headers or BASE_HEADERS)
    if referer:
        hdrs["Referer"] = referer

    last_err = ""
    delay = RETRY_DELAY
    for attempt in range(1 + RETRY_TIMES):
        try:
            r = SESSION.get(url, headers=hdrs, timeout=timeout, allow_redirects=True)
            if r.status_code == 200:
                return r, ""
            return None, f"HTTP {r.status_code}"
        except requests.exceptions.Timeout:
            last_err = "timeout"
        except requests.exceptions.ConnectionError:
            last_err = "connection error"
        except Exception as e:
            return None, str(e)
        if attempt < RETRY_TIMES:
            time.sleep(delay)
            delay *= 2
    return None, last_err


def is_pdf(resp: requests.Response) -> bool:
    ct = resp.headers.get("Content-Type", "")
    return "pdf" in ct.lower()


def pdf_from_response(resp: requests.Response) -> bytes | None:
    if is_pdf(resp) and len(resp.content) > 5_000:
        return resp.content
    return None


def try_download_url(url: str, timeout: int = 30,
                     referer: str = None) -> tuple[bytes | None, str]:
    hdrs = dict(PDF_HEADERS)
    if referer:
        hdrs["Referer"] = referer
    last_err = ""
    delay = RETRY_DELAY
    for attempt in range(1 + RETRY_TIMES):
        try:
            r = SESSION.get(url, headers=hdrs, timeout=timeout, allow_redirects=True)
            # HTTP 204 = No Content — server may require auth redirect; retry as GET
            if r.status_code == 204:
                last_err = "HTTP 204 (no content)"
                break
            if r.status_code != 200:
                return None, f"HTTP {r.status_code}"
            pdf = pdf_from_response(r)
            if pdf:
                return pdf, ""
            ct = r.headers.get("Content-Type", "unknown")
            return None, f"not a PDF (Content-Type: {ct}, size: {len(r.content)})"
        except requests.exceptions.Timeout:
            last_err = "timeout"
        except requests.exceptions.ConnectionError:
            last_err = "connection error"
        except Exception as e:
            return None, str(e)
        if attempt < RETRY_TIMES:
            time.sleep(delay)
            delay *= 2
    return None, last_err


# ── Excel scanning ─────────────────────────────────────────────────────────────

def collect_articles(wb) -> list[dict]:
    articles = []
    for sheet_name in SHEETS_TO_SCAN:
        if sheet_name not in wb.sheetnames:
            print(f"  [WARN] Sheet '{sheet_name}' not found — skipping.")
            continue
        ws = wb[sheet_name]
        header = [cell.value for cell in ws[1]]
        try:
            doi_idx = header.index(DOI_COLUMN)
        except ValueError:
            print(f"  [WARN] Column '{DOI_COLUMN}' not in '{sheet_name}' — skipping.")
            continue
        title_idx = header.index(TITLE_COLUMN) if TITLE_COLUMN in header else None

        yellow_count = green_count = 0
        for row in ws.iter_rows(min_row=2):
            color = row_color(row)
            doi = clean_doi(row[doi_idx].value)
            if not doi:
                continue
            title = str(row[title_idx].value).strip() if title_idx is not None and row[title_idx].value else ""
            entry = {"doi": doi, "title": title}
            if color == YELLOW_COLOR:
                entry["folder"] = sheet_name
                articles.append(entry)
                yellow_count += 1
            elif color in GREEN_COLORS:
                entry["folder"] = FUTURE_FOLDER
                articles.append(entry)
                green_count += 1

        print(f"  '{sheet_name}': {yellow_count} yellow | {green_count} green")
    return articles


# ── Download layers ────────────────────────────────────────────────────────────

def layer_unpaywall(doi: str, **_) -> tuple[bytes | None, str]:
    url = f"https://api.unpaywall.org/v2/{doi}?email={UNPAYWALL_EMAIL}"
    try:
        r, err = fetch(url, timeout=15)
        if not r:
            return None, f"API error: {err}"
        data = r.json()
        if not data.get("is_oa"):
            return None, "not OA"
        locations = data.get("oa_locations") or []
        best = data.get("best_oa_location")
        if best:
            locations = [best] + [l for l in locations if l != best]
        for loc in locations:
            pdf_url = loc.get("url_for_pdf") or loc.get("url")
            if not pdf_url:
                continue
            pdf, reason = try_download_url(pdf_url, referer=f"https://doi.org/{doi}")
            if pdf:
                return pdf, ""
        return None, "OA locations found but no PDF downloadable"
    except Exception as e:
        return None, str(e)


def layer_core(doi: str, **_) -> tuple[bytes | None, str]:
    core_headers = {**BASE_HEADERS, "Authorization": f"Bearer {CORE_API_KEY}"}
    # Try direct DOI lookup first (most precise)
    for url in [
        f"https://api.core.ac.uk/v3/works/doi:{doi}",
        f"https://api.core.ac.uk/v3/search/works?q=doi%3A%22{quote_plus(doi)}%22&limit=1",
    ]:
        try:
            r, err = fetch(url, timeout=15, headers=core_headers)
            if not r:
                continue
            data = r.json()
            # Direct lookup returns object; search returns {results: [...]}
            results = data.get("results") or ([data] if data.get("id") else [])
            if not results:
                continue
            item = results[0]
            candidates = []
            if item.get("downloadUrl"):
                candidates.append(item["downloadUrl"])
            for fu in item.get("fullTextIdentifier") or []:
                candidates.append(fu)
            for fu in item.get("sourceFulltextUrls") or []:
                candidates.append(fu)
            for candidate in candidates:
                if not candidate:
                    continue
                pdf, reason = try_download_url(candidate, referer=f"https://doi.org/{doi}")
                if pdf:
                    return pdf, ""
            if candidates:
                return None, "CORE URLs found but no PDF downloadable"
        except Exception:
            continue
    return None, "not found in CORE"


def layer_semantic_scholar(doi: str, **_) -> tuple[bytes | None, str]:
    url = f"https://api.semanticscholar.org/graph/v1/paper/DOI:{doi}?fields=openAccessPdf"
    try:
        r, err = fetch(url, timeout=15)
        if not r:
            return None, f"API error: {err}"
        data = r.json()
        oa = data.get("openAccessPdf") or {}
        pdf_url = oa.get("url")
        if not pdf_url:
            return None, "no OA PDF in response"
        pdf, reason = try_download_url(pdf_url, referer=f"https://doi.org/{doi}")
        return (pdf, "") if pdf else (None, reason)
    except Exception as e:
        return None, str(e)


def layer_openalex(doi: str, **_) -> tuple[bytes | None, str]:
    url = f"https://api.openalex.org/works/https://doi.org/{doi}"
    try:
        r, err = fetch(url, timeout=15)
        if not r:
            return None, f"API error: {err}"
        data = r.json()
        candidate_urls = []
        oa = data.get("open_access") or {}
        if oa.get("oa_url"):
            candidate_urls.append(oa["oa_url"])
        for loc in data.get("locations") or []:
            pu = loc.get("pdf_url")
            lu = loc.get("landing_page_url")
            if pu:
                candidate_urls.append(pu)
            elif lu and loc.get("is_oa"):
                candidate_urls.append(lu)
        if not candidate_urls:
            return None, "no OA URLs in response"
        for cu in candidate_urls:
            pdf, reason = try_download_url(cu, referer=f"https://doi.org/{doi}")
            if pdf:
                return pdf, ""
        return None, "URLs found but no PDF downloadable"
    except Exception as e:
        return None, str(e)


def layer_pubmed_central(doi: str, **_) -> tuple[bytes | None, str]:
    search_url = (
        f"https://eutils.ncbi.nlm.nih.gov/entrez/eutils/esearch.fcgi"
        f"?db=pmc&term={quote_plus(doi)}[doi]&retmode=json"
    )
    try:
        r, err = fetch(search_url, timeout=15)
        if not r:
            return None, f"API error: {err}"
        ids = r.json().get("esearchresult", {}).get("idlist", [])
        if not ids:
            return None, "not in PMC"
        pmcid = f"PMC{ids[0]}"
        pdf_url = f"https://www.ncbi.nlm.nih.gov/pmc/articles/{pmcid}/pdf/"
        pdf, reason = try_download_url(pdf_url, referer=f"https://doi.org/{doi}")
        return (pdf, "") if pdf else (None, reason)
    except Exception as e:
        return None, str(e)


def layer_europe_pmc(doi: str, **_) -> tuple[bytes | None, str]:
    search_url = (
        f"https://www.ebi.ac.uk/europepmc/webservices/rest/search"
        f"?query=DOI:{doi}&format=json&resultType=core"
    )
    try:
        r, err = fetch(search_url, timeout=15)
        if not r:
            return None, f"API error: {err}"
        results = r.json().get("resultList", {}).get("result", [])
        if not results:
            return None, "not found in Europe PMC"
        pmcid = results[0].get("pmcid")
        if not pmcid:
            return None, "found but no PMCID (not OA)"
        pdf_url = f"https://europepmc.org/articles/{pmcid}/pdf/render"
        pdf, reason = try_download_url(pdf_url, referer=f"https://doi.org/{doi}")
        return (pdf, "") if pdf else (None, reason)
    except Exception as e:
        return None, str(e)


_PDF_PATTERNS = [
    r'<meta[^>]+name=["\']citation_pdf_url["\'][^>]+content=["\']([^"\']+)["\']',
    r'<meta[^>]+content=["\']([^"\']+)["\'][^>]+name=["\']citation_pdf_url["\']',
    r'href=["\']([^"\']*\.pdf[^"\']*)["\']',
    r'data-pdf-url=["\']([^"\']+)["\']',
    r'"pdfUrl"\s*:\s*"([^"]+)"',
    r'"pdf_url"\s*:\s*"([^"]+)"',
]


def find_pdf_url_in_html(html: str, base_url: str) -> str | None:
    for pattern in _PDF_PATTERNS:
        for m in re.findall(pattern, html, re.IGNORECASE):
            if any(x in m.lower() for x in ["thumb", "cover", "suppl", "fig", "icon"]):
                continue
            full = make_absolute(m, base_url)
            if full:
                return full
    return None


def layer_scihub(doi: str, title: str = "", **_) -> tuple[bytes | None, str]:
    last_reason = "no mirrors reachable"

    def _try_scihub_page(url: str) -> tuple[bytes | None, str]:
        resp, err = fetch(url, timeout=30)
        if not resp:
            return None, err
        pdf_url = None
        for pattern in [
            r'<iframe[^>]+src=["\']([^"\']+)["\']',
            r'<embed[^>]+src=["\']([^"\']+)["\']',
        ]:
            for m in re.findall(pattern, resp.text, re.IGNORECASE):
                candidate = make_absolute(m, resp.url)
                if candidate:
                    pdf_url = candidate
                    break
            if pdf_url:
                break
        if not pdf_url:
            pdf_url = find_pdf_url_in_html(resp.text, resp.url)
        if not pdf_url:
            return None, f"no PDF URL found in HTML ({url})"
        pdf, reason = try_download_url(pdf_url, timeout=60, referer=url)
        return (pdf, "") if pdf else (None, f"PDF URL found but {reason}")

    for mirror in SCIHUB_MIRRORS:
        pdf, reason = _try_scihub_page(f"{mirror}/{doi}")
        if pdf:
            return pdf, ""
        last_reason = f"{mirror}/doi: {reason}"

    # Fallback: search by title if available
    if title:
        for mirror in SCIHUB_MIRRORS[:2]:
            search_url = f"{mirror}/?request={quote_plus(title)}"
            pdf, reason = _try_scihub_page(search_url)
            if pdf:
                return pdf, ""
            last_reason = f"{mirror}/title: {reason}"

    return None, last_reason


def ojs_pdf_url(page_url: str, html: str) -> str | None:
    """
    Open Journal Systems uses /article/view/{id}/{galley}
    The PDF download URL is /article/download/{id}/{galley}
    """
    m = re.search(r'/article/view/(\d+)/(\d+)', page_url)
    if m:
        base = page_url[:page_url.index('/article/view/')]
        return f"{base}/article/download/{m.group(1)}/{m.group(2)}"
    # Also handle /index.php/{journal}/article/view pattern
    m2 = re.search(r'(/index\.php/[^/]+/article)/view/(\d+)/(\d+)', page_url)
    if m2:
        base = urlparse(page_url).scheme + "://" + urlparse(page_url).netloc
        return f"{base}{m2.group(1)}/download/{m2.group(2)}/{m2.group(3)}"
    return None


def layer_vpn_html(doi: str, **_) -> tuple[bytes | None, str]:
    doi_url = f"https://doi.org/{doi}"
    resp, err = fetch(doi_url, timeout=30, referer="https://scholar.google.com/")
    if not resp:
        return None, f"doi.org unreachable: {err}"
    if is_pdf(resp):
        pdf = resp.content if len(resp.content) > 5_000 else None
        return (pdf, "") if pdf else (None, "response was PDF but too small")

    final_url = resp.url

    # OJS detection
    ojs_url = ojs_pdf_url(final_url, resp.text)
    if ojs_url:
        pdf, reason = try_download_url(ojs_url, referer=final_url)
        if pdf:
            return pdf, ""

    pdf_url = find_pdf_url_in_html(resp.text, final_url)
    if not pdf_url:
        return None, f"landed on {final_url} — no PDF link in static HTML (may need JS)"
    pdf, reason = try_download_url(pdf_url, referer=final_url)
    return (pdf, "") if pdf else (None, f"PDF URL {pdf_url}: {reason}")


# ── Pipeline ───────────────────────────────────────────────────────────────────

LAYERS = [
    ("Unpaywall",        layer_unpaywall),
    ("CORE",             layer_core),
    ("Semantic Scholar", layer_semantic_scholar),
    ("OpenAlex",         layer_openalex),
    ("PubMed Central",   layer_pubmed_central),
    ("Europe PMC",       layer_europe_pmc),
    ("Sci-Hub",          layer_scihub),
    ("VPN/HTML",         layer_vpn_html),
]


def download_pdf(doi: str, title: str = "") -> tuple[bytes | None, str | None, dict]:
    reasons = {}
    for name, fn in LAYERS:
        try:
            pdf, reason = fn(doi=doi, title=title)
            if pdf:
                return pdf, name, reasons
            reasons[name] = reason or "no PDF"
        except Exception as e:
            reasons[name] = f"exception: {e}"
    return None, None, reasons


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    failed_csv = "failed_downloads.csv"

    if os.path.exists(failed_csv):
        print(f"Found {failed_csv} — running in RETRY mode (only failed DOIs).\n")
        with open(failed_csv, encoding="utf-8") as f:
            reader = csv.DictReader(f)
            articles = [{"doi": row["doi"], "folder": row["folder"], "title": row.get("title", "")}
                        for row in reader]
    else:
        if not os.path.exists(EXCEL_FILE):
            print(f"[ERROR] '{EXCEL_FILE}' not found in current directory.")
            return
        print(f"Reading '{EXCEL_FILE}' ...")
        wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
        print("Scanning sheets ...")
        articles = collect_articles(wb)

    total = len(articles)
    print(f"Total articles to process: {total}\n")

    folders_needed = {a["folder"] for a in articles}
    for folder in folders_needed:
        os.makedirs(folder, exist_ok=True)

    success = fail = skipped = 0
    failed_list = []
    log_rows = []

    for idx, art in enumerate(articles, start=1):
        doi    = art["doi"]
        folder = art["folder"]
        title  = art.get("title", "")
        path   = os.path.join(folder, sanitize_filename(doi) + ".pdf")

        print(f"[{idx}/{total}] {doi}")

        if os.path.exists(path):
            print(f"  [SKIP] Already exists.")
            skipped += 1
            log_rows.append({"doi": doi, "folder": folder, "status": "skip",
                              "source": "", "reasons": ""})
            continue

        pdf_bytes, source, reasons = download_pdf(doi, title=title)

        if pdf_bytes:
            with open(path, "wb") as f:
                f.write(pdf_bytes)
            print(f"  [OK] {source} → {path}")
            success += 1
            log_rows.append({"doi": doi, "folder": folder, "status": "ok",
                              "source": source, "reasons": ""})
        else:
            reason_str = " | ".join(f"{k}: {v}" for k, v in reasons.items())
            print(f"  [FAIL] {reason_str}")
            failed_list.append({"doi": doi, "folder": folder, "title": title,
                                 "doi_url": f"https://doi.org/{doi}",
                                 "reasons": reason_str})
            log_rows.append({"doi": doi, "folder": folder, "status": "fail",
                              "source": "", "reasons": reason_str})
            fail += 1

        time.sleep(DELAY_SECONDS)

    print(f"\n{'='*60}")
    print(f"Done.  OK: {success}  |  Failed: {fail}  |  Skipped: {skipped}  |  Total: {total}")
    for folder in sorted(folders_needed):
        if os.path.isdir(folder):
            count = len([f for f in os.listdir(folder) if f.endswith(".pdf")])
            print(f"  {folder}/  → {count} PDFs")

    if failed_list:
        with open(failed_csv, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=["doi", "folder", "title", "doi_url", "reasons"])
            writer.writeheader()
            writer.writerows(failed_list)
        print(f"\nFailed DOIs → {failed_csv} ({len(failed_list)} entries)")
    elif os.path.exists(failed_csv):
        os.remove(failed_csv)
        print(f"\nAll retries succeeded — {failed_csv} removed.")

    with open("download_log.csv", "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=["doi", "folder", "status", "source", "reasons"])
        writer.writeheader()
        writer.writerows(log_rows)
    print(f"Full log → download_log.csv")


if __name__ == "__main__":
    main()
