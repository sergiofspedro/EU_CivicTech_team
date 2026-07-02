"""
build_new_batch_excel.py
Creates an Excel overview for the "New batch" folder, split into the two
NotebookLM notebooks you divided it into ("Selected 1" and "Selected 2").

The PDFs listed in SELECTED_2_FILES go to the "Selected 2" sheet.
Every other PDF found in NEW_BATCH_FOLDER goes to the "Selected 1" sheet.
Metadata is looked up from the original Scopus Excel by DOI (sanitized
filename), same logic as build_batch_excel.py.

Usage:
  python build_new_batch_excel.py
"""

import os
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ─── CONFIGURATION ─────────────────────────────────────────────────────────────
SCOPUS_EXCEL    = "Nets4Dem_ScopusReview_Final_GA_12 June.xlsx"
SCOPUS_SHEETS   = ["HIGH relevance", "MEDIUM screening"]
NEW_BATCH_FOLDER = r"C:\Users\Administrator\Downloads\Livro\New batch"
OUTPUT_FILE     = "new_batch_overview.xlsx"

# PDFs (filename without .pdf) that you added to NotebookLM "Selected 2".
# Everything else found in NEW_BATCH_FOLDER is treated as "Selected 1".
SELECTED_2_FILES = [
    "10.1163_18763375-01001007",
    "10.1007_s10784-008-9073-7",
    "10.1002_csr.1940",
    "10.1177_00936502211016162",
    "10.1145_3428502.3428614",
    "10.1017_CBO9780511762642.022",
    "10.1080_19460171.2017.1393440",
    "10.1080_0703633042000222367",
    "10.1016_j.polsoc.2009.02.001",
    "10.1080_17513057.2012.719632",
    "10.1111_jcms.13411",
    "10.1080_17448689.2015.1069524",
    "10.1016_j.futures.2012.07.007",
    "10.1177_0888325411401380",
]
# ───────────────────────────────────────────────────────────────────────────────


def sanitize_filename(doi: str) -> str:
    return re.sub(r'[\\/:*?"<>|]', "_", doi)


def build_doi_lookup(excel_file: str) -> dict:
    wb = openpyxl.load_workbook(excel_file, data_only=True)
    lookup = {}

    for sheet_name in SCOPUS_SHEETS:
        if sheet_name not in wb.sheetnames:
            print(f"  [WARN] Sheet '{sheet_name}' not found in source Excel")
            continue
        ws = wb[sheet_name]
        headers = [cell.value for cell in ws[1]]

        doi_idx = next((i for i, h in enumerate(headers)
                        if h and str(h).strip().upper() == "DOI"), None)
        if doi_idx is None:
            print(f"  [WARN] No DOI column in '{sheet_name}'")
            continue

        for row in ws.iter_rows(min_row=2, values_only=True):
            raw_doi = row[doi_idx]
            if not raw_doi:
                continue
            doi = str(raw_doi).strip()
            for prefix in ("https://doi.org/", "http://doi.org/", "doi.org/", "DOI:", "doi:"):
                if doi.lower().startswith(prefix.lower()):
                    doi = doi[len(prefix):]
            doi = doi.strip()
            if not doi or doi.lower() == "nan":
                continue

            key = sanitize_filename(doi)
            record = {"_source_sheet": sheet_name}
            for col_name, value in zip(headers, row):
                if col_name:
                    record[str(col_name)] = value
            lookup[key] = record

    print(f"  Loaded {len(lookup)} articles from source Excel.")
    return lookup


def get_pdf_keys(folder_path: str) -> list[str]:
    if not os.path.isdir(folder_path):
        print(f"  [WARN] Folder not found: {folder_path}")
        return []
    return sorted(f[:-4] for f in os.listdir(folder_path) if f.lower().endswith(".pdf"))


def write_sheet(ws, headers: list, rows: list):
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)

    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for row in rows:
        ws.append(row)

    for col_idx, header in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = len(str(header))
        for row_cells in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row_cells:
                if cell.value:
                    max_len = max(max_len, min(len(str(cell.value)), 60))
        ws.column_dimensions[col_letter].width = max_len + 2

    ws.freeze_panes = "A2"


def main():
    if not os.path.exists(SCOPUS_EXCEL):
        print(f"[ERROR] Source Excel not found: {SCOPUS_EXCEL}")
        return

    print(f"Reading source Excel: {SCOPUS_EXCEL}")
    lookup = build_doi_lookup(SCOPUS_EXCEL)

    all_cols = []
    seen = set()
    for record in lookup.values():
        for k in record:
            if k != "_source_sheet" and k not in seen:
                all_cols.append(k)
                seen.add(k)

    output_headers = ["Source sheet"] + all_cols + ["PDF filename", "Matched"]

    print(f"\nScanning '{NEW_BATCH_FOLDER}' ...")
    all_keys = get_pdf_keys(NEW_BATCH_FOLDER)
    print(f"  {len(all_keys)} PDFs found in New batch folder")

    selected_2_set = set(SELECTED_2_FILES)
    not_found_in_folder = selected_2_set - set(all_keys)
    if not_found_in_folder:
        print(f"\n[WARN] {len(not_found_in_folder)} files listed for Selected 2 "
              f"were not found in the folder:")
        for f in sorted(not_found_in_folder):
            print(f"    {f}")

    groups = {
        "Selected 2": [k for k in all_keys if k in selected_2_set],
        "Selected 1": [k for k in all_keys if k not in selected_2_set],
    }

    out_wb = openpyxl.Workbook()
    out_wb.remove(out_wb.active)

    summary_rows = []

    # Write Selected 1 first, then Selected 2, for natural reading order
    for sheet_name in ["Selected 1", "Selected 2"]:
        keys = groups[sheet_name]
        print(f"\nProcessing '{sheet_name}': {len(keys)} PDFs")

        rows = []
        matched = unmatched = 0
        for key in keys:
            record = lookup.get(key)
            if record:
                row = [record.get("_source_sheet", "")]
                for col in all_cols:
                    row.append(record.get(col, ""))
                row.append(key + ".pdf")
                row.append("YES")
                matched += 1
            else:
                row = [""] * (len(all_cols) + 1)
                row.append(key + ".pdf")
                row.append("NO - not found in Excel")
                unmatched += 1
            rows.append(row)

        ws = out_wb.create_sheet(title=sheet_name)
        write_sheet(ws, output_headers, rows)
        print(f"  Matched: {matched}  |  Unmatched: {unmatched}")
        summary_rows.append((sheet_name, NEW_BATCH_FOLDER, len(keys), matched, unmatched))

    ws_sum = out_wb.create_sheet(title="Summary", index=0)
    sum_headers = ["Notebook", "Folder", "Total PDFs", "Matched", "Unmatched"]
    write_sheet(ws_sum, sum_headers, summary_rows)

    out_wb.save(OUTPUT_FILE)
    print(f"\n{'='*60}")
    print(f"Output saved to: {os.path.abspath(OUTPUT_FILE)}")
    for name, folder, total, matched, unmatched in summary_rows:
        print(f"  {name}: {total} PDFs | {matched} matched | {unmatched} unmatched")


if __name__ == "__main__":
    main()
